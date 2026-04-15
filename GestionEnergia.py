import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

# ==========================================
# MODELOS DE DATOS
# ==========================================

class Variables:
    def __init__(self, nom_mae="", nom_inv="", nom_ce="", e_punta=0.0, e_llano=0.0, e_valle=0.0,
                 pot_contratada=0.0, e_pot_punta=0.0, e_pot_valle=0.0, e_excedentes=0.0, iva=0.0):
        self.nom_mae = nom_mae
        self.nom_inv = nom_inv
        self.nom_ce = nom_ce
        self.e_punta = float(e_punta)
        self.e_llano = float(e_llano)
        self.e_valle = float(e_valle)
        self.pot_contratada = float(pot_contratada)
        self.e_pot_punta = float(e_pot_punta)
        self.e_pot_valle = float(e_pot_valle)
        self.e_excedentes = float(e_excedentes)
        self.iva = float(iva)

    def total_potencia(self):
        return (self.pot_contratada * 1 * self.e_pot_punta) + (self.pot_contratada * 1 * self.e_pot_valle)


class DatosComun:
    def __init__(self, nombre, fecha=None):
        self.nombre = nombre
        # Si no se pasa fecha, asigna la fecha actual en formato dd/mm/yyyy
        self.fecha = fecha if fecha else datetime.now().strftime("%d/%m/%Y")


class MaquinaAerotermia(DatosComun):
    def __init__(self, nombre, fecha, kwh_total, kalef_kw, euros, agua_kw=0.0):
        super().__init__(nombre, fecha)
        self.kwh_total = float(kwh_total)
        self.kalef_kw = float(kalef_kw)
        self.euros = float(euros)
        self.agua_kw = float(agua_kw)


class Inversor(DatosComun):
    def __init__(self, nombre, fecha, vertida, consumo_real, consumo_total, solar):
        super().__init__(nombre, fecha)
        self.vertida = float(vertida)
        self.consumo_real = float(consumo_real)
        self.consumo_total = float(consumo_total)
        self.solar = float(solar)


class CompaniaElectrica(DatosComun):
    def __init__(self, nombre, fecha, punta, llano, valle, vertida, kwh_total, variables: Variables):
        super().__init__(nombre, fecha)
        self.punta = float(punta)
        self.llano = float(llano)
        self.valle = float(valle)
        self.vertida = float(vertida)
        self.kwh_total = float(kwh_total)
        self.euros = self.calculo_de_euros(variables)

    def calculo_de_euros(self, v: Variables):
        coste_energia = (self.punta * v.e_punta) + (self.valle * v.e_valle) + (self.llano * v.e_llano)
        beneficio_vertida = self.vertida * v.e_excedentes
        total_sin_iva = (coste_energia + v.total_potencia()) - beneficio_vertida
        return total_sin_iva * v.iva


# ==========================================
# MANEJO DE DATOS (EQUIVALENTE A APACHE POI)
# ==========================================

class ManejoDatos:
    def __init__(self):
        self.hoja_registros = "Registros"
        # En openpyxl los índices empiezan en 1
        self.fila_cabecera_variables = 35 
        self.fila_variables = 36          

        # Definición de Colores (Hexadecimal sin el '#', formato aRGB)
        self.color_mae = PatternFill(start_color="FFFFB6C1", end_color="FFFFB6C1", fill_type="solid") # Rosa pastel
        self.color_inv = PatternFill(start_color="FFADD8E6", end_color="FFADD8E6", fill_type="solid") # Azul pastel
        self.color_ce = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")  # Verde pastel
        self.fuente_negrita = Font(bold=True)

    def pedir_nombre_excel(self):
        nombre = input("\nIntroduce el nombre del archivo Excel a utilizar (sin añadir .xlsx): ").strip()
        return f"{nombre}.xlsx"

    def aplicar_estilo_fila(self, hoja, fila, is_header=False):
        # Aplica los colores por columnas según el bloque (MAE, INV, CE)
        if is_header:
            hoja.cell(row=fila, column=1).font = self.fuente_negrita
            
        for col in range(2, 6): # MAE
            hoja.cell(row=fila, column=col).fill = self.color_mae
            if is_header: hoja.cell(row=fila, column=col).font = self.fuente_negrita
        for col in range(6, 10): # INV
            hoja.cell(row=fila, column=col).fill = self.color_inv
            if is_header: hoja.cell(row=fila, column=col).font = self.fuente_negrita
        for col in range(10, 16): # CE
            hoja.cell(row=fila, column=col).fill = self.color_ce
            if is_header: hoja.cell(row=fila, column=col).font = self.fuente_negrita

    def crear_tabla(self):
        archivo_excel = self.pedir_nombre_excel()

        if not os.path.exists(archivo_excel):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = self.hoja_registros

            # Cabeceras Registros (Fila 1)
            cabeceras_reg = ["Fecha", "MAE_KwhTotal", "MAE_AguaKw", "MAE_KalefKw", "MAE_Euros",
                             "INV_Vertida", "INV_ConsumoReal", "INV_ConsumoTotal", "INV_Solar",
                             "CE_Punta", "CE_Llano", "CE_Valle", "CE_Vertida", "CE_KwhTotal", "CE_Euros"]
            sheet.append(cabeceras_reg)
            self.aplicar_estilo_fila(sheet, 1, is_header=True)

            # Cabeceras Variables (Fila 35)
            cabeceras_vars = ["NomMAE", "NomINV", "NomCE", "EurosPunta", "EurosLlano", "EurosValle",
                              "PotContratada", "EurosPotPunta", "EurosPotValle", "EurosExcedentes", "IVA"]
            for col, val in enumerate(cabeceras_vars, start=1):
                cell = sheet.cell(row=self.fila_cabecera_variables, column=col, value=val)
                cell.font = self.fuente_negrita
                if col == 1: cell.fill = self.color_mae
                elif col == 2: cell.fill = self.color_inv
                else: cell.fill = self.color_ce

            wb.save(archivo_excel)
            print(f"Archivo Excel '{archivo_excel}' creado exitosamente con colores.")
        else:
            print(f"El archivo Excel '{archivo_excel}' ya existe.")

    def insertar_variables(self):
        archivo_excel = self.pedir_nombre_excel()
        if not os.path.exists(archivo_excel):
            print("El archivo no existe. Crea la tabla primero.")
            return

        print("--- Inserción de Variables ---")
        v = Variables(
            nom_mae=input("Nombre Máquina Aerotermia: "),
            nom_inv=input("Nombre Inversor: "),
            nom_ce=input("Nombre Compañía Eléctrica: "),
            e_punta=input("Euros Punta: "),
            e_llano=input("Euros Llano: "),
            e_valle=input("Euros Valle: "),
            pot_contratada=input("Potencia Contratada: "),
            e_pot_punta=input("Euros Potencia Punta: "),
            e_pot_valle=input("Euros Potencia Valle: "),
            e_excedentes=input("Euros Excedentes: "),
            iva=input("IVA (ej. 0.21): ")
        )

        wb = openpyxl.load_workbook(archivo_excel)
        sheet = wb[self.hoja_registros]

        datos = [v.nom_mae, v.nom_inv, v.nom_ce, v.e_punta, v.e_llano, v.e_valle,
                 v.pot_contratada, v.e_pot_punta, v.e_pot_valle, v.e_excedentes, v.iva]

        for col, val in enumerate(datos, start=1):
            cell = sheet.cell(row=self.fila_variables, column=col, value=val)
            if col == 1: cell.fill = self.color_mae
            elif col == 2: cell.fill = self.color_inv
            else: cell.fill = self.color_ce

        wb.save(archivo_excel)
        print("Variables guardadas y coloreadas correctamente.")

    def eliminar_variables(self):
        archivo_excel = self.pedir_nombre_excel()
        try:
            wb = openpyxl.load_workbook(archivo_excel)
            sheet = wb[self.hoja_registros]
            # Borrar el contenido de la fila de variables
            for col in range(1, 12):
                sheet.cell(row=self.fila_variables, column=col).value = None
            wb.save(archivo_excel)
            print("Variables eliminadas.")
        except Exception as e:
            print(f"Error al eliminar variables: {e}")

    def obtener_variables_del_excel(self, archivo_excel):
        try:
            wb = openpyxl.load_workbook(archivo_excel)
            sheet = wb[self.hoja_registros]
            if sheet.cell(row=self.fila_variables, column=1).value is None:
                return None
            
            return Variables(
                *[sheet.cell(row=self.fila_variables, column=c).value for c in range(1, 12)]
            )
        except Exception:
            return None

    def buscar_fila_vacia(self, sheet):
        for r in range(2, self.fila_cabecera_variables):
            if sheet.cell(row=r, column=1).value is None:
                return r
        return -1

    def escribir_datos_registro(self, sheet, fila, fecha, vars):
        print(f"--- Máquina Aerotermia ({vars.nom_mae}) ---")
        mae = MaquinaAerotermia(vars.nom_mae, fecha, 
                                kwh_total=input("Kwh Total: "),
                                agua_kw=input("Agua Kw: "),
                                kalef_kw=input("Kalefaccion Kw: "),
                                euros=input("Euros: "))

        print(f"--- Inversor ({vars.nom_inv}) ---")
        inv = Inversor(vars.nom_inv, fecha,
                       vertida=input("Vertida: "),
                       consumo_real=input("Consumo Real: "),
                       consumo_total=input("Consumo Total: "),
                       solar=input("Solar: "))

        print(f"--- Compañía Eléctrica ({vars.nom_ce}) ---")
        ce = CompaniaElectrica(vars.nom_ce, fecha,
                               punta=input("Punta: "),
                               llano=input("Llano: "),
                               valle=input("Valle: "),
                               vertida=input("Vertida: "),
                               kwh_total=input("KwhTotal: "),
                               variables=vars)

        datos = [fecha, mae.kwh_total, mae.agua_kw, mae.kalef_kw, mae.euros,
                 inv.vertida, inv.consumo_real, inv.consumo_total, inv.solar,
                 ce.punta, ce.llano, ce.valle, ce.vertida, ce.kwh_total, ce.euros]

        for col, val in enumerate(datos, start=1):
            sheet.cell(row=fila, column=col, value=val)
        
        # Aplicamos negrita a la fecha y colores a las secciones
        self.aplicar_estilo_fila(sheet, fila, is_header=False)
        sheet.cell(row=fila, column=1).font = self.fuente_negrita

    def insertar_elemento(self):
        archivo_excel = self.pedir_nombre_excel()
        vars = self.obtener_variables_del_excel(archivo_excel)
        if not vars:
            print("ERROR: No hay variables instanciadas. Ejecuta 'Insertar Variables' primero.")
            return

        wb = openpyxl.load_workbook(archivo_excel)
        sheet = wb[self.hoja_registros]
        fila_destino = self.buscar_fila_vacia(sheet)

        if fila_destino == -1:
            print("ERROR: No hay espacio libre.")
            return

        fecha = input("Introduce la fecha (dd/MM/yyyy): ")
        self.escribir_datos_registro(sheet, fila_destino, fecha, vars)
        
        wb.save(archivo_excel)
        print(f"Elemento guardado correctamente en la fila {fila_destino}.")

    def eliminar_elemento(self):
        archivo_excel = self.pedir_nombre_excel()
        fecha_eliminar = input("Introduce la fecha del elemento a eliminar (dd/MM/yyyy): ")

        wb = openpyxl.load_workbook(archivo_excel)
        sheet = wb[self.hoja_registros]
        encontrado = False

        for r in range(2, self.fila_cabecera_variables):
            if sheet.cell(row=r, column=1).value == fecha_eliminar:
                # openpyxl borra la fila y desplaza automáticamente las de abajo hacia arriba!
                sheet.delete_rows(r, 1)
                # Como se desplaza todo hacia arriba, debemos insertar una fila vacía justo antes
                # de las variables para que las variables no se suban de posición
                sheet.insert_rows(self.fila_cabecera_variables - 1, 1)
                encontrado = True
                print("Elemento eliminado.")
                break

        if encontrado:
            wb.save(archivo_excel)
        else:
            print("No se ha encontrado elemento con esa fecha.")

    def modificar_elemento(self):
        archivo_excel = self.pedir_nombre_excel()
        vars = self.obtener_variables_del_excel(archivo_excel)
        if not vars: return

        fecha_modificar = input("Introduce la fecha del elemento a modificar (dd/MM/yyyy): ")
        wb = openpyxl.load_workbook(archivo_excel)
        sheet = wb[self.hoja_registros]

        for r in range(2, self.fila_cabecera_variables):
            if sheet.cell(row=r, column=1).value == fecha_modificar:
                print("Elemento encontrado. Introduce los nuevos datos:")
                self.escribir_datos_registro(sheet, r, fecha_modificar, vars)
                wb.save(archivo_excel)
                print("Elemento modificado exitosamente.")
                return
        
        print("No se ha encontrado elemento con esa fecha.")

    def mostrar_elemento(self):
        archivo_excel = self.pedir_nombre_excel()
        fecha_buscada = input("Introduce la fecha exacta del registro a consultar (dd/MM/yyyy): ")

        try:
            wb = openpyxl.load_workbook(archivo_excel, data_only=True)
            sheet = wb[self.hoja_registros]

            for r in range(2, self.fila_cabecera_variables):
                if sheet.cell(row=r, column=1).value == fecha_buscada:
                    print("==================================================")
                    print(f"        DATOS REGISTRADOS EL: {fecha_buscada}")
                    print("==================================================")
                    print("\n[ MÁQUINA AEROTERMIA ]")
                    print(f"  - Kwh Total:       {sheet.cell(r, 2).value}")
                    print(f"  - Coste estimado:  {sheet.cell(r, 5).value} €")

                    print("\n[ INVERSOR SOLAR ]")
                    print(f"  - Consumo Total:   {sheet.cell(r, 8).value} kWh")
                    print(f"  - Energía Solar:   {sheet.cell(r, 9).value} kWh")

                    print("\n[ COMPAÑÍA ELÉCTRICA ]")
                    print(f"  - Kwh Total Red:   {sheet.cell(r, 14).value} kWh")
                    print(f"  - Coste Total:     {sheet.cell(r, 15).value} €")
                    print("==================================================\n")
                    return
            
            print(f"No se ha encontrado ningún registro para la fecha ({fecha_buscada}).")
        except Exception as e:
            print(f"Error al leer el archivo: {e}")


# ==========================================
# MENÚ E INICIO
# ==========================================

def menu_inicio():
    manejo_datos = ManejoDatos()
    
    while True:
        print("\n===== GESTIÓN DE ENERGÍA =====")
        print("1. Crear nuevo archivo Excel")
        print("2. Insertar Variables")
        print("3. Eliminar Variables")
        print("4. Añadir registro (Elemento)")
        print("5. Mostrar registros (por fecha)")
        print("6. Modificar registro (Elemento)")
        print("7. Borrar registro (Elemento)")
        print("8. Salir")
        print("==============================")
        
        opcion = input("Elige una opción: ")

        try:
            if opcion == '1': manejo_datos.crear_tabla()
            elif opcion == '2': manejo_datos.insertar_variables()
            elif opcion == '3': manejo_datos.eliminar_variables()
            elif opcion == '4': manejo_datos.insertar_elemento()
            elif opcion == '5': manejo_datos.mostrar_elemento()
            elif opcion == '6': manejo_datos.modificar_elemento()
            elif opcion == '7': manejo_datos.eliminar_elemento()
            elif opcion == '8':
                print("Saliendo del programa... ¡Hasta pronto!")
                break
            else:
                print("Opción no válida. Por favor, elige un número del 1 al 8.")
        except ValueError:
            print("Error numérico. Asegúrate de introducir números donde se piden.")

if __name__ == "__main__":
    menu_inicio()